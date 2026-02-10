"""
EU FUNDING RADAR — Bilbao Misión Climática
============================================
Consulta la API pública de la Comisión Europea (SEDIA),
filtra convocatorias relevantes para el Ayuntamiento de Bilbao
y genera un informe HTML + Excel con fichas + alertas por email.

Uso:  python eu_funding_radar.py
Requisitos:  pip install openpyxl
"""

import json
import os
import re
import sys
import smtplib
import time
import urllib.request
import urllib.parse
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timezone
from pathlib import Path

# Intentar importar openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("⚠️  openpyxl no instalado. Ejecuta: pip install openpyxl")
    print("   El Excel no se generará, pero el HTML sí.\n")

# ──────────────────────────────────────────────
# CONFIGURACIÓN
# ──────────────────────────────────────────────

CONFIG = {
    "keywords": [
        "climate neutral cities",
        "smart cities climate",
        "sustainable urban mobility",
        "energy efficiency buildings cities",
        "nature-based solutions urban",
        "circular economy cities",
        "clean energy transition local",
        "green infrastructure urban",
        "climate adaptation cities",
        "zero emission transport urban",
        "biodiversity urban nature",
        "net zero cities mission",
        "LIFE climate action local",
        "renewable energy communities",
        "urban resilience climate",
        "URBACT action planning",
        "INTERREG climate cooperation",
        "building renovation energy",
        "Innovation Fund decarbonisation",
        "Innovation Fund energy storage",
        "CEF transport sustainable",
        "CEF energy infrastructure",
        "Digital Europe smart cities AI",
        "digital twin urban",
        "hydrogen cities",
        "heat pump district heating",
        "electric bus public transport",
        "waste management circular",
        "water management urban climate",
        "flood risk adaptation",
        "coastal resilience cities",
        "green deal local authorities",
        "just transition fund",
        "cohesion policy climate",
        "ERDF sustainable urban development",
        "social climate fund",
        "mission ocean coastal cities",
        "mission soil land restoration",
        "new european bauhaus",
        "european urban initiative innovative actions",
    ],

    "email_to": os.environ.get("EMAIL_TO", ""),
    "email_from": os.environ.get("EMAIL_FROM", ""),
    "smtp_server": os.environ.get("SMTP_SERVER", "smtp.gmail.com"),
    "smtp_port": int(os.environ.get("SMTP_PORT") or "587"),
    "smtp_user": os.environ.get("SMTP_USER", ""),
    "smtp_pass": os.environ.get("SMTP_PASS", ""),

    "seen_file": "seen_calls.json",
    "output_file": "resultados_convocatorias.json",
    "output_html": "resultados_convocatorias.html",
    "output_excel": "resultados_convocatorias.xlsx",
}

# Relevancia por keywords para Bilbao
BILBAO_RELEVANCE = {
    "climate neutral cities": ("MUY ALTA", "Bilbao es ciudad Mission de neutralidad climática 2030. Acceso prioritario vía NetZeroCities."),
    "net zero cities mission": ("MUY ALTA", "Directamente vinculado al Climate City Contract de Bilbao."),
    "smart cities climate": ("ALTA", "Encaja con estrategia Smart City de Bilbao (AS Fabrik, Zorrotzaurre)."),
    "sustainable urban mobility": ("MUY ALTA", "Bilbao es nodo TEN-T. Metro, tranvía, BizkaiBus. Proyecto SuperIsla."),
    "energy efficiency buildings": ("MUY ALTA", "Parque edificatorio antiguo (Casco Viejo, Bilbao La Vieja). Plan rehabilitación activo."),
    "building renovation energy": ("MUY ALTA", "Alineado con Estrategia de Rehabilitación Energética municipal."),
    "nature-based solutions": ("ALTA", "Proyectos de renaturalización en Zorrotzaurre y Abandoibarra. Plan Verde municipal."),
    "green infrastructure": ("ALTA", "Anillo Verde de Bilbao y corredores ecológicos en desarrollo."),
    "circular economy": ("ALTA", "Bilbao tiene Plan de Economía Circular y proyectos en Mercabilbao."),
    "clean energy transition": ("ALTA", "Comunidades energéticas en desarrollo (Otxarkoaga, Txurdinaga)."),
    "renewable energy communities": ("ALTA", "Estrategia energética municipal incluye comunidades energéticas locales."),
    "climate adaptation": ("ALTA", "Plan de Adaptación al Cambio Climático de Bilbao vigente."),
    "zero emission transport": ("ALTA", "Flota municipal en electrificación. Zona bajas emisiones en estudio."),
    "biodiversity urban": ("MEDIA", "Urdaibai cerca. Biodiversidad urbana en parques y ría."),
    "LIFE climate action": ("ALTA", "LIFE es programa clave para acción climática local. Bilbao elegible."),
    "URBACT action planning": ("MUY ALTA", "Bilbao tiene experiencia previa en redes URBACT. Ideal para intercambio."),
    "INTERREG climate cooperation": ("ALTA", "Elegible SUDOE e INTERREG Atlantic. Socios potenciales: Burdeos, Oporto."),
    "Innovation Fund": ("MEDIA", "Aplicable a Petronor/Repsol (Muskiz) y Puerto de Bilbao."),
    "CEF transport": ("MUY ALTA", "Bilbao es nodo TEN-T. Infraestructuras de transporte elegibles."),
    "CEF energy": ("MEDIA", "Infraestructuras energéticas del País Vasco potencialmente elegibles."),
    "Digital Europe": ("ALTA", "Experiencia en gemelo digital urbano. Complementa estrategia Smart City."),
    "digital twin": ("ALTA", "Proyecto BIM/GIS de Bilbao. Complementa estrategia digital."),
    "hydrogen": ("MEDIA", "Corredor Vasco del Hidrógeno. Petronor y Puerto de Bilbao involucrados."),
    "heat pump": ("MEDIA", "Aplicable a rehabilitación de edificios municipales."),
    "electric bus": ("ALTA", "BizkaiBus en proceso de electrificación de flota."),
    "waste management": ("ALTA", "Garbiker y gestión de residuos municipal. Plan de residuos activo."),
    "water management": ("MEDIA", "Consorcio de Aguas Bilbao Bizkaia. Gestión de ría y pluviales."),
    "flood risk": ("ALTA", "Bilbao tiene historial de inundaciones (ría del Nervión). Planes activos."),
    "coastal resilience": ("MEDIA", "Proximidad costera. Impacto indirecto vía estuario del Nervión."),
    "green deal local": ("ALTA", "Pacto Verde aplicable a nivel local. Bilbao comprometida."),
    "just transition": ("MEDIA", "Margen izquierda con pasado industrial. Potencial reconversión."),
    "cohesion policy": ("MEDIA", "País Vasco es región en transición. Fondos FEDER disponibles."),
    "ERDF sustainable urban": ("ALTA", "Bilbao es elegible para EDUSI/FEDER urbano sostenible."),
    "social climate fund": ("MEDIA", "Aplicable a pobreza energética y movilidad sostenible asequible."),
    "mission ocean coastal": ("MEDIA", "Estuario del Nervión y proximidad al Cantábrico."),
    "mission soil": ("MEDIA", "Suelos contaminados industriales en Zorrotzaurre y Margen Izquierda."),
    "new european bauhaus": ("ALTA", "NEB combina sostenibilidad + diseño + inclusion. Ideal para regeneracion urbana Bilbao (Zorrotzaurre, Casco Viejo)."),
    "european urban initiative": ("MUY ALTA", "EUI-IA financia proyectos urbanos innovadores. Bilbao como ciudad Mission es candidata ideal."),
}


# ──────────────────────────────────────────────
# API DE LA COMISIÓN EUROPEA (SEDIA)
# ──────────────────────────────────────────────

BASE_URL = "https://api.tech.ec.europa.eu/search-api/prod/rest/search"

def search_eu_api(keyword, page_size=50):
    params = urllib.parse.urlencode({
        "apiKey": "SEDIA",
        "text": keyword,
        "pageSize": str(page_size),
        "pageNumber": "1",
    })
    url = f"{BASE_URL}?{params}"
    try:
        req = urllib.request.Request(
            url,
            data=b"",
            headers={"User-Agent": "Mozilla/5.0 (EU-Funding-Radar-Bilbao/2.0)"},
        )
        with urllib.request.urlopen(req, timeout=30) as response:
            return json.loads(response.read().decode("utf-8"))
    except Exception as e:
        print(f"  ⚠️  Error: {e}")
        return None


def get_relevance_for_call(call):
    """Determina la relevancia para Bilbao basándose en keywords."""
    best_level = "INFO"
    best_note = "Revisar relevancia para el Ayuntamiento de Bilbao."
    level_order = {"MUY ALTA": 4, "ALTA": 3, "MEDIA": 2, "BAJA": 1, "INFO": 0}

    call_text = f"{call.get('id','')} {call.get('title','')} {call.get('description','')}".lower()

    for keyword_fragment, (level, note) in BILBAO_RELEVANCE.items():
        if keyword_fragment.lower() in call_text:
            if level_order.get(level, 0) > level_order.get(best_level, 0):
                best_level = level
                best_note = note
    return best_level, best_note


def parse_results(api_response):
    calls = []
    if not api_response or "results" not in api_response:
        return calls

    for item in api_response.get("results", []):
        try:
            ref = item.get("reference", "")
            url = item.get("url", "")
            content = item.get("content", "")
            title = item.get("title", "")
            summary = item.get("summary", "")
            metadata = item.get("metadata", {})

            # ─── TOPIC ID ───
            # Preferir metadata.identifier
            topic_id = ""
            if metadata.get("identifier"):
                topic_id = metadata["identifier"][0]
            elif url:
                parts = url.rstrip("/").split("/")
                if parts:
                    topic_id = parts[-1]
            if topic_id.endswith(".json"):
                topic_id = topic_id[:-5]
            if not topic_id:
                match = re.search(r'(HORIZON-[A-Z0-9\-]+|LIFE-[A-Z0-9\-]+|CEF-[A-Z0-9\-]+|DIGITAL-[A-Z0-9\-]+|INTERREG-[A-Z0-9\-]+|INNOVFUND-[A-Z0-9\-]+)', ref)
                if match:
                    topic_id = match.group(1)
                else:
                    topic_id = ref[:60] if ref else ""
            if not topic_id:
                continue

            # ─── TITULO ───
            # Preferir metadata.title sobre el campo title raiz (que suele ser null)
            clean_title = ""
            if metadata.get("title"):
                clean_title = re.sub(r'<[^>]+>', '', str(metadata["title"][0])).strip()
            if not clean_title or clean_title == "None":
                clean_title = re.sub(r'<[^>]+>', '', str(title or "")).strip()
            if not clean_title or clean_title == "None":
                clean_title = re.sub(r'<[^>]+>', '', str(summary or "")).strip()
            if not clean_title or clean_title == "None":
                clean_title = topic_id

            # ─── DESCRIPCION ───
            clean_summary = re.sub(r'<[^>]+>', '', str(summary or "")).strip()[:500]
            clean_content = re.sub(r'<[^>]+>', '', str(content or "")).strip()[:300]
            description = clean_summary or clean_content

            # Intentar obtener descripcion mas rica de descriptionByte
            if metadata.get("descriptionByte"):
                raw_desc = re.sub(r'<[^>]+>', '', str(metadata["descriptionByte"][0])).strip()
                if len(raw_desc) > len(description):
                    description = raw_desc[:600]

            # ─── ESTADO ───
            status = "Unknown"
            # Extraer de metadata.actions que tiene el status real
            if metadata.get("actions"):
                try:
                    actions_data = json.loads(metadata["actions"][0])
                    if isinstance(actions_data, list) and actions_data:
                        action_status = actions_data[0].get("status", {}).get("abbreviation", "")
                        if action_status:
                            status = action_status  # "Closed", "Open", "Forthcoming"
                except (json.JSONDecodeError, IndexError, KeyError):
                    pass
            # Fallback: buscar en texto
            if status == "Unknown":
                text_lower = (clean_content + clean_summary + str(metadata.get("sortStatus", ""))).lower()
                if "forthcoming" in text_lower or "upcoming" in text_lower:
                    status = "Forthcoming"
                elif "open" in text_lower:
                    status = "Open"
                elif "closed" in text_lower:
                    status = "Closed"

            # ─── DEADLINE ───
            deadline = ""
            if metadata.get("deadlineDate"):
                try:
                    raw_date = metadata["deadlineDate"][0]
                    # Formato: "2023-04-27T00:00:00.000+0000"
                    date_part = raw_date[:10]  # "2023-04-27"
                    dt = datetime.strptime(date_part, "%Y-%m-%d")
                    deadline = dt.strftime("%d/%m/%Y")
                except (ValueError, IndexError):
                    deadline = raw_date[:10] if raw_date else ""
            # Fallback: extraer del actions
            if not deadline and metadata.get("actions"):
                try:
                    actions_data = json.loads(metadata["actions"][0])
                    if isinstance(actions_data, list) and actions_data:
                        dd = actions_data[0].get("deadlineDates", [])
                        if dd:
                            dt = datetime.strptime(dd[0], "%Y-%m-%d")
                            deadline = dt.strftime("%d/%m/%Y")
                except:
                    pass

            # ─── PROGRAMA ───
            programme = ""
            tid_upper = topic_id.upper()
            if "HORIZON" in tid_upper:
                programme = "Horizon Europe"
            elif "LIFE" in tid_upper:
                programme = "LIFE"
            elif "CEF" in tid_upper:
                programme = "CEF"
            elif "DIGITAL" in tid_upper:
                programme = "Digital Europe"
            elif "INTERREG" in tid_upper:
                programme = "INTERREG"
            elif "INNOVFUND" in tid_upper:
                programme = "Innovation Fund"
            elif "URBACT" in tid_upper:
                programme = "URBACT"

            # ─── TIPO DE ACCION ───
            action_type = ""
            if metadata.get("typesOfAction"):
                action_type = metadata["typesOfAction"][0].replace("HORIZON ", "")

            # ─── PRESUPUESTO ───
            budget = ""
            if metadata.get("budgetOverview"):
                try:
                    bo = json.loads(metadata["budgetOverview"][0])
                    for topic_key, actions in bo.get("budgetTopicActionMap", {}).items():
                        for a in actions:
                            min_c = a.get("minContribution", 0)
                            max_c = a.get("maxContribution", 0)
                            if max_c:
                                if min_c == max_c:
                                    budget = f"{max_c:,.0f}"
                                else:
                                    budget = f"{min_c:,.0f} - {max_c:,.0f}"
                                break
                        if budget:
                            break
                except:
                    pass

            # ─── CALL ID ───
            call_id = ""
            if metadata.get("callIdentifier"):
                call_id = metadata["callIdentifier"][0]

            # ─── URL ───
            # Preferir metadata.url sobre url raiz
            final_url = ""
            if metadata.get("url"):
                final_url = metadata["url"][0]
            elif url:
                final_url = url
            if not final_url or "topic-details" not in final_url:
                if "calls-for-proposals" not in final_url and "competitive-calls" not in final_url:
                    # Construir URL manualmente
                    if topic_id:
                        final_url = f"https://ec.europa.eu/info/funding-tenders/opportunities/portal/screen/opportunities/topic-details/{topic_id}"
                    else:
                        continue
            if final_url.endswith(".json"):
                continue

            # ─── TAGS ───
            tags = metadata.get("tags", [])

            call_data = {
                "id": topic_id,
                "title": clean_title,
                "status": status,
                "programme": programme,
                "deadline": deadline,
                "description": description,
                "url": final_url,
                "action_type": action_type,
                "budget": budget,
                "call_id": call_id,
                "tags": ", ".join(tags) if tags else "",
                "fetched_at": datetime.now(timezone.utc).isoformat(),
            }

            # Anadir relevancia para Bilbao
            rel_level, rel_note = get_relevance_for_call(call_data)
            call_data["relevance_level"] = rel_level
            call_data["relevance_note"] = rel_note

            calls.append(call_data)
        except Exception:
            continue
    return calls


# ──────────────────────────────────────────────
# LÓGICA PRINCIPAL
# ──────────────────────────────────────────────

def load_seen():
    path = Path(CONFIG["seen_file"])
    if path.exists():
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            return {}
    return {}


def save_seen(seen):
    with open(CONFIG["seen_file"], "w", encoding="utf-8") as f:
        json.dump(seen, f, ensure_ascii=False, indent=2)


def fetch_all_calls():
    all_calls = {}
    total = len(CONFIG["keywords"])
    today = datetime.now(timezone.utc)

    print(f"\n🇪🇺 EU FUNDING RADAR — Bilbao Misión Climática")
    print(f"{'='*50}")
    print(f"📅 {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print(f"🔍 Buscando en {total} categorías...\n")

    for i, keyword in enumerate(CONFIG["keywords"], 1):
        print(f"  [{i}/{total}] {keyword}...", end=" ", flush=True)
        response = search_eu_api(keyword)
        if response:
            total_hits = response.get("totalResults", 0)
            calls = parse_results(response)
            new = 0
            for call in calls:
                if call["id"] not in all_calls:
                    all_calls[call["id"]] = call
                    new += 1
            print(f"✓ {total_hits} hits, {new} convocatorias nuevas")
        else:
            print("✗ error")

    print(f"\n📊 Total convocatorias encontradas: {len(all_calls)}")

    # ─── FILTRADO ESTRICTO POR FECHAS ───
    filtered = {}
    skipped_closed = 0
    skipped_old = 0

    for k, v in all_calls.items():
        status = v.get("status", "Unknown")
        deadline_str = v.get("deadline", "")

        # 1. Saltar las explicitamente cerradas
        if status == "Closed":
            skipped_closed += 1
            continue

        # 2. Si tiene deadline, comprobar si ya ha pasado
        if deadline_str:
            try:
                deadline_date = datetime.strptime(deadline_str, "%d/%m/%Y").replace(tzinfo=timezone.utc)
                if deadline_date < today:
                    # Deadline ya pasado -> cerrada aunque la API diga otra cosa
                    skipped_old += 1
                    continue
            except ValueError:
                pass  # No se pudo parsear, mantener

        # 3. Si no tiene deadline, comprobar el año del topic ID
        # Convocatorias de 2023 o 2024 sin deadline probablemente ya estan cerradas
        if not deadline_str:
            topic_id = v.get("id", "")
            if re.search(r'202[0-4]', topic_id):
                skipped_old += 1
                continue

        # 4. Corregir status basado en deadline
        if deadline_str and status != "Forthcoming":
            try:
                deadline_date = datetime.strptime(deadline_str, "%d/%m/%Y").replace(tzinfo=timezone.utc)
                if deadline_date > today:
                    v["status"] = "Open"
            except ValueError:
                pass

        filtered[k] = v

    print(f"📊 Descartadas cerradas: {skipped_closed}")
    print(f"📊 Descartadas con deadline pasado: {skipped_old}")
    print(f"📊 Convocatorias vigentes: {len(filtered)}")
    return filtered


# ──────────────────────────────────────────────
# GENERACIÓN EXCEL
# ──────────────────────────────────────────────

def generate_excel(all_calls, new_calls):
    if not HAS_OPENPYXL:
        print("⚠️  Saltando Excel (openpyxl no instalado)")
        return

    wb = Workbook()

    # Colors
    DARK_BLUE = "1B2A4A"
    HEADER_BLUE = "0057B7"
    LIGHT_BLUE = "E8F0FE"
    LIGHT_GREEN = "E6F4EA"
    LIGHT_RED = "FCE8E6"
    LIGHT_YELLOW = "FFF8E1"
    WHITE = "FFFFFF"
    BORDER_GRAY = "D1D5DB"

    thin_border = Border(
        left=Side(style='thin', color=BORDER_GRAY),
        right=Side(style='thin', color=BORDER_GRAY),
        top=Side(style='thin', color=BORDER_GRAY),
        bottom=Side(style='thin', color=BORDER_GRAY),
    )
    header_font = Font(name='Leelawadee UI', bold=True, color=WHITE, size=11)
    header_fill = PatternFill('solid', fgColor=HEADER_BLUE)
    title_font = Font(name='Leelawadee UI', bold=True, size=16, color=DARK_BLUE)
    subtitle_font = Font(name='Leelawadee UI', size=11, color="6B7280")

    # Sort: source group (EU, ES, Euskadi) -> status -> relevance -> deadline
    def excel_sort_key(x):
        source = x.get("source", "EU")
        if source == "BDNS":
            src_order = 1
        elif source == "KontratazioA":
            src_order = 2
        else:
            src_order = 0
        rel_order = {"MUY ALTA": 0, "ALTA": 1, "MEDIA": 2}.get(x.get("relevance_level", ""), 3)
        status_order = {"Open": 0, "Forthcoming": 1}.get(x.get("status", ""), 2)
        return (src_order, status_order, rel_order, x.get("deadline", "99/99/9999"))

    calls_sorted = sorted(all_calls.values(), key=excel_sort_key)

    # ─── SHEET 1: RESUMEN ───
    ws = wb.active
    ws.title = "Resumen"

    ws.merge_cells('A1:M1')
    ws['A1'] = "EU FUNDING RADAR — AYUNTAMIENTO DE BILBAO"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 40

    ws.merge_cells('A2:M2')
    ws['A2'] = f"Mision Climatica - Neutralidad 2030 - Actualizado: {datetime.now().strftime('%d/%m/%Y %H:%M')} - {len(all_calls)} convocatorias - {len(new_calls)} nuevas"
    ws['A2'].font = subtitle_font
    ws.row_dimensions[2].height = 22

    headers = [
        ("Fuente", 12),
        ("ID Convocatoria", 28),
        ("Titulo", 55),
        ("Programa", 18),
        ("Estado", 14),
        ("Deadline", 15),
        ("Presupuesto", 22),
        ("Tipo Accion", 20),
        ("Relevancia Bilbao", 18),
        ("Nota Relevancia", 40),
        ("Descripcion", 50),
        ("Nueva?", 10),
        ("Link", 12),
    ]

    for col, (name, width) in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[4].height = 32

    status_fills = {
        "Open": PatternFill('solid', fgColor=LIGHT_GREEN),
        "Forthcoming": PatternFill('solid', fgColor=LIGHT_BLUE),
        "Closed": PatternFill('solid', fgColor=LIGHT_RED),
    }
    relevance_fills = {
        "MUY ALTA": PatternFill('solid', fgColor="DCFCE7"),
        "ALTA": PatternFill('solid', fgColor=LIGHT_BLUE),
        "MEDIA": PatternFill('solid', fgColor=LIGHT_YELLOW),
    }
    source_fills = {
        "EU": PatternFill('solid', fgColor="DBEAFE"),
        "BDNS": PatternFill('solid', fgColor="FEF3C7"),
        "KontratazioA": PatternFill('solid', fgColor="D1FAE5"),
    }
    source_fonts = {
        "EU": Font(name='Leelawadee UI', size=10, bold=True, color="1E40AF"),
        "BDNS": Font(name='Leelawadee UI', size=10, bold=True, color="92400E"),
        "KontratazioA": Font(name='Leelawadee UI', size=10, bold=True, color="065F46"),
    }
    source_labels = {
        "EU": "🇪🇺 Europa",
        "BDNS": "🇪🇸 España",
        "KontratazioA": "🟢 Euskadi",
    }

    # Write section headers + data rows
    current_source = None
    row_num = 4  # header row

    for call in calls_sorted:
        source = call.get("source", "EU")
        is_new = call["id"] in new_calls

        # Insert source group separator
        if source != current_source:
            current_source = source
            row_num += 1
            label = source_labels.get(source, source)
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=13)
            sep_cell = ws.cell(row=row_num, column=1, value=label)
            sep_colors = {"EU": "1E40AF", "BDNS": "B45309", "KontratazioA": "065F46"}
            sep_cell.font = Font(name='Leelawadee UI', size=12, bold=True, color="FFFFFF")
            sep_cell.fill = PatternFill('solid', fgColor=sep_colors.get(source, "1E40AF"))
            sep_cell.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[row_num].height = 28

        row_num += 1
        ws.row_dimensions[row_num].height = 40

        source_label = {"EU": "Europa", "BDNS": "España", "KontratazioA": "Euskadi"}.get(source, source)

        values = [
            source_label,
            call["id"],
            call["title"],
            call["programme"],
            call["status"],
            call.get("deadline", ""),
            call.get("budget", ""),
            call.get("action_type", ""),
            call.get("relevance_level", "INFO"),
            call.get("relevance_note", ""),
            call["description"][:200],
            "NUEVA" if is_new else "",
            "Ver",
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = Font(name='Leelawadee UI', size=10)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = thin_border

            if col == 1:  # Fuente
                cell.fill = source_fills.get(source, PatternFill())
                cell.font = source_fonts.get(source, Font(name='Leelawadee UI', size=10))
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col == 5:  # Status
                cell.fill = status_fills.get(val, PatternFill())
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='Leelawadee UI', size=10, bold=True)
            elif col == 9:  # Relevance
                cell.fill = relevance_fills.get(val, PatternFill())
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='Leelawadee UI', size=10, bold=True)
            elif col == 12 and is_new:  # New badge
                cell.font = Font(name='Leelawadee UI', size=10, bold=True, color="DC2626")
                cell.fill = PatternFill('solid', fgColor="FEF2F2")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col == 13:  # Link
                cell.font = Font(name='Leelawadee UI', size=10, color="0057B7", underline='single')
                cell.hyperlink = call["url"]
                cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.auto_filter.ref = f"A4:M{row_num}"
    ws.freeze_panes = "A5"

    # ─── SHEET 2: FICHAS DETALLADAS ───
    ws2 = wb.create_sheet("Fichas Detalladas")

    row = 1
    for i, call in enumerate(calls_sorted):
        # Header
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws2.cell(row=row, column=1, value=f"FICHA {i+1}: {call['id']}")
        cell.font = Font(name='Leelawadee UI', bold=True, size=12, color=WHITE)
        cell.fill = header_fill
        for c in range(1, 7):
            ws2.cell(row=row, column=c).fill = header_fill
            ws2.cell(row=row, column=c).border = thin_border
        ws2.row_dimensions[row].height = 30
        row += 1

        source = call.get("source", "EU")
        source_label = {"EU": "🇪🇺 Europa", "BDNS": "🇪🇸 España", "KontratazioA": "🟢 Euskadi"}.get(source, source)
        fields = [
            ("Fuente", source_label),
            ("Titulo", call["title"]),
            ("ID", call["id"]),
            ("Programa", call["programme"]),
            ("Estado", call["status"]),
            ("Deadline", call.get("deadline", "No disponible")),
            ("Presupuesto (EUR)", call.get("budget", "No disponible") or "No disponible"),
            ("Tipo de Accion", call.get("action_type", "No disponible") or "No disponible"),
            ("Call ID", call.get("call_id", "") or ""),
            ("Descripcion", call["description"]),
            ("Tags", call.get("tags", "") or ""),
            ("Relevancia Bilbao", f"{call.get('relevance_level', 'INFO')} — {call.get('relevance_note', '')}"),
            ("Enlace al portal", call["url"]),
        ]

        for label, value in fields:
            ws2.cell(row=row, column=1, value=label)
            ws2.cell(row=row, column=1).font = Font(name='Leelawadee UI', bold=True, size=10, color=DARK_BLUE)
            ws2.cell(row=row, column=1).fill = PatternFill('solid', fgColor="F1F5F9")
            ws2.cell(row=row, column=1).alignment = Alignment(vertical='top')
            ws2.cell(row=row, column=1).border = thin_border

            ws2.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            cell = ws2.cell(row=row, column=2, value=value)
            cell.font = Font(name='Leelawadee UI', size=10)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border
            for c in range(2, 7):
                ws2.cell(row=row, column=c).border = thin_border

            if label == "Relevancia Bilbao":
                if "MUY ALTA" in str(value):
                    cell.font = Font(name='Leelawadee UI', size=10, bold=True, color="166534")
                elif "ALTA" in str(value):
                    cell.font = Font(name='Leelawadee UI', size=10, bold=True, color="1E40AF")
            elif label == "Enlace al portal":
                cell.font = Font(name='Leelawadee UI', size=10, color="0057B7", underline='single')
                cell.hyperlink = value

            ws2.row_dimensions[row].height = 20 if len(str(value)) < 80 else 45
            row += 1
        row += 1

    ws2.column_dimensions['A'].width = 22
    for c in 'BCDEF':
        ws2.column_dimensions[c].width = 20

    # ─── SHEET 3: SEGUIMIENTO ───
    ws3 = wb.create_sheet("Seguimiento")

    track_headers = [
        ("ID Convocatoria", 30),
        ("Titulo corto", 40),
        ("Programa", 16),
        ("Estado", 14),
        ("Deadline", 18),
        ("Relevancia", 14),
        ("Responsable", 20),
        ("Estado interno", 22),
        ("Socios identificados", 30),
        ("Notas / Proximos pasos", 40),
    ]

    for col, (name, width) in enumerate(track_headers, 1):
        cell = ws3.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = PatternFill('solid', fgColor="7C3AED")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws3.column_dimensions[get_column_letter(col)].width = width
    ws3.row_dimensions[1].height = 32

    for i, call in enumerate(calls_sorted):
        row = 2 + i
        ws3.cell(row=row, column=1, value=call["id"]).font = Font(name='Leelawadee UI', size=10)
        ws3.cell(row=row, column=2, value=call["title"][:60]).font = Font(name='Leelawadee UI', size=10)
        ws3.cell(row=row, column=3, value=call["programme"]).font = Font(name='Leelawadee UI', size=10)
        ws3.cell(row=row, column=4, value=call["status"]).font = Font(name='Leelawadee UI', size=10)
        ws3.cell(row=row, column=5, value=call.get("deadline", "")).font = Font(name='Leelawadee UI', size=10)
        ws3.cell(row=row, column=6, value=call.get("relevance_level", "")).font = Font(name='Leelawadee UI', size=10, bold=True)
        ws3.cell(row=row, column=7, value="").font = Font(name='Leelawadee UI', size=10)
        ws3.cell(row=row, column=8, value="Por revisar").font = Font(name='Leelawadee UI', size=10)
        ws3.cell(row=row, column=9, value="").font = Font(name='Leelawadee UI', size=10)
        ws3.cell(row=row, column=10, value="").font = Font(name='Leelawadee UI', size=10)
        for col in range(1, 11):
            ws3.cell(row=row, column=col).border = thin_border
            ws3.cell(row=row, column=col).alignment = Alignment(vertical='center', wrap_text=True)
        ws3.row_dimensions[row].height = 28

    ws3.auto_filter.ref = f"A1:J{1 + len(calls_sorted)}"
    ws3.freeze_panes = "A2"

    # ─── SHEET 4: RECURSOS ───
    ws4 = wb.create_sheet("Recursos")

    resources = [
        # --- EUROPA ---
        ("SECCION: EUROPA", "", ""),
        ("Portal EU Funding & Tenders", "https://ec.europa.eu/info/funding-tenders/opportunities/portal/screen/opportunities/calls-for-proposals", "Portal principal de convocatorias europeas"),
        ("NetZeroCities (Mission Cities)", "https://netzerocities.eu", "Plataforma ciudades Mission. Bilbao es miembro."),
        ("URBACT", "https://urbact.eu/calls-for-proposals", "Redes de ciudades europeas"),
        ("INTERREG SUDOE", "https://interreg-sudoe.eu", "Cooperacion ES-FR-PT"),
        ("INTERREG Atlantic", "https://www.atlanticarea.eu", "Cooperacion Arco Atlantico"),
        ("LIFE Programme (CINEA)", "https://cinea.ec.europa.eu/programmes/life_en", "Programa medioambiental y climatico"),
        ("Innovation Fund", "https://climate.ec.europa.eu/eu-action/eu-funding-climate-action/innovation-fund_en", "Descarbonizacion industrial"),
        ("CEF Transport", "https://cinea.ec.europa.eu/programmes/connecting-europe-facility/transport_en", "Connecting Europe - Transporte"),
        ("Digital Europe", "https://digital-strategy.ec.europa.eu/en/activities/digital-programme", "IA, datos, ciberseguridad"),
        ("EIT Climate-KIC", "https://www.climate-kic.org/programmes/", "Innovacion climatica - Convocatorias propias"),
        ("EIT Urban Mobility", "https://www.eiturbanmobility.eu/calls/", "Movilidad urbana - Convocatorias propias"),
        ("EIT InnoEnergy", "https://www.innoenergy.com/", "Energia sostenible - Convocatorias propias"),
        ("New European Bauhaus", "https://new-european-bauhaus.europa.eu/get-involved/funding-opportunities_en", "Sostenibilidad, estetica e inclusion - Convocatorias propias"),
        ("EUI - Innovative Actions", "https://www.urban-initiative.eu/calls-for-proposals", "European Urban Initiative - Acciones Innovadoras para ciudades"),
        ("BEI - Programa ELENA", "https://www.eib.org/en/products/advising/elena/index.htm", "Asistencia tecnica BEI para eficiencia energetica"),
        ("EIB Circular City Centre", "https://advisory.eib.org/about/circular-city-centre.htm", "Centro asesoramiento BEI para economia circular urbana"),
        ("C40 Cities", "https://www.c40.org/", "Red global ciudades por el clima - Recursos y buenas practicas"),
        ("TED - Licitaciones EU", "https://ted.europa.eu/en/advanced-search", "Tenders Electronic Daily - Licitaciones publicas europeas"),
        ("CORDIS", "https://cordis.europa.eu", "Base de datos de proyectos financiados"),
        # --- ESPANA ---
        ("SECCION: ESPANA", "", ""),
        ("IDAE Convocatorias abiertas", "https://ayudasenergiaidae.es/programas-ayudas-abiertas/", "Eficiencia energetica, renovables, MOVES, PRTR"),
        ("IDAE Catalogo de ayudas", "https://www.idae.es/ayudas-y-financiacion/catalogo-de-ayudas", "Todas las ayudas IDAE disponibles"),
        ("CDTI Convocatorias", "https://www.cdti.es/convocatorias", "I+D+i, Compra Publica Innovadora"),
        ("Fundacion Biodiversidad", "https://fundacion-biodiversidad.es/convocatorias/", "Renaturalizacion, biodiversidad, empleo verde"),
        ("MITECO Convocatorias", "https://www.miteco.gob.es/es/ministerio/servicios/ayudas-subvenciones/", "Transicion ecologica, reto demografico"),
        ("BDNS Subvenciones", "https://www.pap.hacienda.gob.es/bdnstrans/GE/es/convocatorias", "Base de Datos Nacional de Subvenciones - TODAS"),
        ("Red Innpulso", "https://www.redinnpulso.es/", "Red de ciudades de ciencia e innovacion"),
        ("FEMP Fondos Europeos", "https://femp-fondos-europa.es/convocatorias/", "Federacion Municipios - Convocatorias para EELL"),
        # --- EUSKADI / BIZKAIA ---
        ("SECCION: EUSKADI / BIZKAIA", "", ""),
        ("EVE Programa de ayudas", "https://www.eve.eus/programa-de-ayudas/", "Autoconsumo, renovables, eficiencia, vehiculos"),
        ("Ihobe Subvenciones", "https://www.ihobe.eus/subvenciones", "Ecoinnovacion, economia circular, clima"),
        ("SPRI Programas", "https://www.spri.eus/es/ayudas/", "Digitalizacion, industria, smart cities"),
        ("Diputacion Bizkaia Subvenciones", "https://www.bizkaia.eus/es/subvenciones", "Ayudas departamentales para municipios"),
        ("Gobierno Vasco Ayudas", "https://www.euskadi.eus/ayudas-subvenciones-702/web01-tramite/es/", "Portal completo ayudas Gobierno Vasco"),
        ("Udalsarea 2030", "https://www.udalsarea21.net/", "Red vasca municipios sostenibles"),
        ("Barrixe - Rehabilitacion Bilbao", "https://barrixe.com/", "Oficina rehabilitacion y regeneracion urbana Bilbao"),
        ("KontratazioA - Licitaciones Euskadi", "https://www.contratacion.euskadi.eus/webkpe00-kpeperfi/es/ac70cPublicidadWar/busquedaAnuncios?locale=es", "Plataforma contratacion publica sector publico vasco"),
    ]

    ws4.merge_cells('A1:C1')
    ws4['A1'] = "RECURSOS Y ENLACES UTILES"
    ws4['A1'].font = Font(name='Leelawadee UI', bold=True, size=14, color=DARK_BLUE)
    ws4.row_dimensions[1].height = 35

    for col, (name, width) in enumerate([("Recurso", 35), ("Enlace", 75), ("Descripcion", 50)], 1):
        cell = ws4.cell(row=3, column=col, value=name)
        cell.font = header_font
        cell.fill = PatternFill('solid', fgColor="059669")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws4.column_dimensions[get_column_letter(col)].width = width

    for i, (name, url, desc) in enumerate(resources):
        row = 4 + i
        if name.startswith("SECCION:"):
            # Section header
            ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            cell = ws4.cell(row=row, column=1, value=name.replace("SECCION: ", ""))
            cell.font = Font(name='Leelawadee UI', bold=True, size=11, color="FFFFFF")
            section_colors = {"EUROPA": "1E40AF", "ESPANA": "92400E", "EUSKADI / BIZKAIA": "065F46"}
            color = "333333"
            for k, v in section_colors.items():
                if k in name:
                    color = v
            cell.fill = PatternFill('solid', fgColor=color)
            for c in range(1, 4):
                ws4.cell(row=row, column=c).fill = PatternFill('solid', fgColor=color)
                ws4.cell(row=row, column=c).border = thin_border
            ws4.row_dimensions[row].height = 28
        else:
            ws4.cell(row=row, column=1, value=name).font = Font(name='Leelawadee UI', bold=True, size=10)
            if url:
                link_cell = ws4.cell(row=row, column=2, value=url)
                link_cell.font = Font(name='Leelawadee UI', size=10, color="0057B7", underline='single')
                link_cell.hyperlink = url
            ws4.cell(row=row, column=3, value=desc).font = Font(name='Leelawadee UI', size=10)
            for col in range(1, 4):
                ws4.cell(row=row, column=col).border = thin_border
                ws4.cell(row=row, column=col).alignment = Alignment(vertical='center', wrap_text=True)
            ws4.row_dimensions[row].height = 28

    wb.save(CONFIG["output_excel"])
    print(f"📊 Excel generado: {CONFIG['output_excel']}")


# ──────────────────────────────────────────────
# INFORME HTML (actualizado con enlace a Excel)
# ──────────────────────────────────────────────


def generate_html(all_calls, new_calls):
    eu_calls = [c for c in all_calls.values() if c.get("source") not in ("BDNS", "KontratazioA")]
    es_calls = [c for c in all_calls.values() if c.get("source") == "BDNS"]
    eus_calls = [c for c in all_calls.values() if c.get("source") == "KontratazioA"]

    def sort_key(x):
        rel_order = {"MUY ALTA": 0, "ALTA": 1, "MEDIA": 2}
        status_order = {"Open": 0, "Forthcoming": 1}
        return (status_order.get(x["status"], 2), rel_order.get(x.get("relevance_level",""), 3), x.get("deadline","99/99/9999"))

    eu_calls.sort(key=sort_key); es_calls.sort(key=sort_key); eus_calls.sort(key=sort_key)
    new_list = sorted([c for c in all_calls.values() if c["id"] in new_calls], key=sort_key)
    muy_alta = sum(1 for c in all_calls.values() if c.get("relevance_level") == "MUY ALTA")

    def make_row(call, show_new=True):
        is_new = call["id"] in new_calls
        new_badge = ' <span class="badge-new">NUEVA</span>' if is_new and show_new else ""
        s = call["status"]
        badge = '<span class="badge-open">Abierta</span>' if s == "Open" else '<span class="badge-forth">Próxima</span>' if s == "Forthcoming" else '<span class="badge-info">Info</span>'
        source = call.get("source", "EU")
        if source == "BDNS":
            src_badge, src_tag = '<span class="badge-src-es">España</span>', "es"
        elif source == "KontratazioA":
            src_badge, src_tag = '<span class="badge-src-eus">Euskadi</span>', "eus"
        else:
            src_badge, src_tag = '<span class="badge-src-eu">Europa</span>', "eu"
        prog = call.get("programme", "")
        prog_badge = f'<span class="badge-prog">{prog[:30]}</span>' if prog else ""
        rel = call.get("relevance_level", "")
        rcls = {"MUY ALTA":"badge-rel-muy","ALTA":"badge-rel-alta","MEDIA":"badge-rel-media"}.get(rel,"")
        rel_badge = f'<span class="{rcls}">{rel}</span>' if rcls else ""
        desc = call["description"][:160]
        rn = call.get("relevance_note","")
        rh = f'<div class="rel-note">💡 {rn}</div>' if rn else ""
        bu = call.get("budget","")
        bh = f'<div class="budget">💰 {bu}</div>' if bu else ""
        dl = call.get("deadline","") or "—"
        at = call.get("action_type","")
        tb = '<span class="badge-type-lic">Licitación</span>' if "Licitacion" in at else '<span class="badge-type-ayuda">Ayuda</span>' if ("Ayuda" in at or "Subvencion" in at) else ""
        return f'''
        <tr class="call-row" data-source="{src_tag}" data-rel="{rel}" data-new="{'1' if is_new else '0'}">
            <td class="cell-main">
                <div class="call-title">{call['title'][:130]}{new_badge}</div>
                <div class="call-meta">{src_badge}{badge}{tb}{prog_badge}{rel_badge}</div>
                <div class="call-desc">{desc}</div>{rh}{bh}
            </td>
            <td class="cell-deadline">{dl}</td>
            <td class="cell-link"><a href="{call['url']}" target="_blank" class="link-ver">Ver →</a></td>
        </tr>'''

    all_sorted = sorted(all_calls.values(), key=sort_key)
    all_rows = "".join(make_row(c) for c in all_sorted)

    if new_list:
        new_rows = "".join(make_row(c, show_new=False) for c in new_list)
        new_section = f'''
        <div class="new-alert" id="new-section">
            <div class="new-alert-title">🆕 {len(new_list)} convocatoria{"s" if len(new_list)>1 else ""} nueva{"s" if len(new_list)>1 else ""}</div>
            <table class="calls-table"><thead><tr class="table-head"><th>Convocatoria</th><th class="th-dl">Deadline</th><th class="th-lk"></th></tr></thead>
            <tbody>{new_rows}</tbody></table></div>'''
    else:
        new_section = '<div class="no-new-alert">✅ Sin novedades desde la última ejecución</div>'

    html = f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Funding Radar — Bilbao</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:opsz,wght@9..40,400;9..40,500;9..40,700&family=JetBrains+Mono:wght@500&display=swap" rel="stylesheet">
<style>
:root{{--bg:#F0F2F5;--card:#FFF;--bdr:#E4E7EC;--tx:#1A1D26;--tx2:#6B7280;--tx3:#9CA3AF;--eu:#1E40AF;--eubg:#EFF6FF;--eubd:#BFDBFE;--es:#B45309;--esbg:#FFFBEB;--esbd:#FCD34D;--eus:#047857;--eusbg:#ECFDF5;--eusbd:#6EE7B7;--red:#DC2626;--grn:#059669;--r:10px}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'DM Sans',sans-serif;background:var(--bg);color:var(--tx);font-size:14px}}
.ctn{{max-width:1060px;margin:0 auto;padding:16px}}
.hdr{{display:flex;align-items:center;justify-content:space-between;padding:18px 22px;background:var(--card);border-radius:var(--r);border:1px solid var(--bdr);margin-bottom:12px}}
.hdr-l{{display:flex;align-items:center;gap:12px}}
.hdr-ico{{width:40px;height:40px;border-radius:10px;background:linear-gradient(135deg,#0057B7,#FFD700);display:flex;align-items:center;justify-content:center;font-size:16px;color:white;font-weight:800}}
.hdr h1{{font-size:19px;font-weight:700;letter-spacing:-0.03em}}
.hdr-sub{{font-size:11px;color:var(--tx3);font-weight:500}}
.hdr-dt{{font-family:'JetBrains Mono',monospace;font-size:11px;color:var(--tx3);background:var(--bg);padding:4px 10px;border-radius:6px}}

.sts{{display:grid;grid-template-columns:repeat(6,1fr);gap:8px;margin-bottom:12px}}
.st{{background:var(--card);border:1px solid var(--bdr);border-radius:var(--r);padding:10px;text-align:center}}
.st-n{{font-size:26px;font-weight:800;letter-spacing:-0.03em;line-height:1}}
.st-l{{font-size:9px;color:var(--tx3);text-transform:uppercase;font-weight:700;letter-spacing:0.06em;margin-top:3px}}

.tabs{{display:flex;background:var(--card);border:1px solid var(--bdr);border-radius:var(--r);padding:4px;margin-bottom:12px}}
.tab{{flex:1;padding:9px 0;text-align:center;border-radius:7px;font-size:12px;font-weight:600;cursor:pointer;border:none;background:transparent;color:var(--tx2);transition:all .15s;font-family:inherit}}
.tab:hover{{background:var(--bg)}}
.tab.active{{color:white}}
.tab.active.t-all{{background:var(--tx)}}
.tab.active.t-eu{{background:var(--eu)}}
.tab.active.t-es{{background:var(--es)}}
.tab.active.t-eus{{background:var(--eus)}}
.tab.active.t-new{{background:var(--red)}}
.tc{{font-family:'JetBrains Mono',monospace;font-size:10px;opacity:.7;margin-left:3px}}

.tbar{{display:flex;gap:8px;margin-bottom:12px;align-items:center;flex-wrap:wrap}}
.srch{{flex:1;min-width:180px;padding:8px 12px;border:1px solid var(--bdr);border-radius:8px;font-size:13px;font-family:inherit;outline:none;background:var(--card)}}
.srch:focus{{border-color:var(--eu);box-shadow:0 0 0 3px rgba(30,64,175,.08)}}
.fb{{padding:7px 12px;border-radius:7px;border:1px solid var(--bdr);background:var(--card);font-size:11px;font-weight:600;cursor:pointer;font-family:inherit;color:var(--tx2)}}
.fb:hover{{background:var(--bg)}}.fb.active{{background:var(--tx);color:white;border-color:var(--tx)}}

.xlbar{{background:var(--eubg);border:1px solid var(--eubd);border-radius:var(--r);padding:8px 14px;margin-bottom:12px;font-size:12px;display:flex;align-items:center;gap:8px}}
.xlbar a{{color:var(--eu);font-weight:700;text-decoration:none}}.xlbar a:hover{{text-decoration:underline}}

.new-alert{{background:#FEF2F2;border:2px solid #FECACA;border-radius:var(--r);padding:14px;margin-bottom:14px}}
.new-alert-title{{font-size:14px;font-weight:700;color:var(--red);margin-bottom:8px}}
.no-new-alert{{background:var(--eusbg);border:1px solid var(--eusbd);border-radius:var(--r);padding:10px 14px;margin-bottom:12px;color:var(--eus);font-weight:600;font-size:13px}}

.calls-table{{width:100%;background:var(--card);border:1px solid var(--bdr);border-radius:var(--r);border-collapse:collapse;overflow:hidden}}
.table-head{{background:#F8FAFC;border-bottom:2px solid var(--bdr)}}
.table-head th{{padding:7px 12px;text-align:left;font-size:9px;color:var(--tx3);font-weight:700;text-transform:uppercase;letter-spacing:.06em}}
.th-dl{{width:85px}}.th-lk{{width:45px}}

.call-row{{border-bottom:1px solid #F3F4F6;transition:background .1s}}
.call-row:hover{{background:#FAFBFD}}
.call-row[data-new="1"]{{background:#FFFBEB}}

.cell-main{{padding:9px 12px}}
.cell-deadline{{padding:9px 12px;font-size:11px;color:#334155;font-weight:600;vertical-align:top;font-family:'JetBrains Mono',monospace;white-space:nowrap}}
.cell-link{{padding:9px 12px;text-align:right;vertical-align:top}}

.call-title{{font-weight:600;font-size:12.5px;line-height:1.35;margin-bottom:4px}}
.call-meta{{display:flex;flex-wrap:wrap;gap:3px;margin-bottom:3px}}
.call-desc{{font-size:11px;color:var(--tx2)}}
.rel-note{{font-size:10px;color:var(--grn);margin-top:2px}}
.budget{{font-size:10px;color:var(--es);margin-top:2px}}

.badge-new{{color:#fff;background:var(--red);padding:1px 6px;border-radius:3px;font-size:9px;font-weight:800;margin-left:4px}}
.badge-open{{color:#065F46;background:#D1FAE5;padding:1px 7px;border-radius:3px;font-size:9px;font-weight:700}}
.badge-forth{{color:var(--eu);background:var(--eubg);padding:1px 7px;border-radius:3px;font-size:9px;font-weight:700}}
.badge-info{{color:var(--tx3);background:#F3F4F6;padding:1px 7px;border-radius:3px;font-size:9px}}
.badge-prog{{color:#6D28D9;background:#F5F3FF;padding:1px 5px;border-radius:3px;font-size:9px;font-weight:600;max-width:180px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;display:inline-block;vertical-align:middle}}
.badge-rel-muy{{color:#065F46;background:#D1FAE5;padding:1px 6px;border-radius:3px;font-size:9px;font-weight:700}}
.badge-rel-alta{{color:var(--eu);background:var(--eubg);padding:1px 6px;border-radius:3px;font-size:9px;font-weight:700}}
.badge-rel-media{{color:var(--es);background:var(--esbg);padding:1px 6px;border-radius:3px;font-size:9px;font-weight:600}}
.badge-src-eu{{color:#fff;background:var(--eu);padding:1px 7px;border-radius:3px;font-size:9px;font-weight:700}}
.badge-src-es{{color:#fff;background:var(--es);padding:1px 7px;border-radius:3px;font-size:9px;font-weight:700}}
.badge-src-eus{{color:#fff;background:var(--eus);padding:1px 7px;border-radius:3px;font-size:9px;font-weight:700}}
.badge-type-lic{{color:#7C2D12;background:#FED7AA;padding:1px 5px;border-radius:3px;font-size:9px;font-weight:600}}
.badge-type-ayuda{{color:#1E3A5F;background:#BAE6FD;padding:1px 5px;border-radius:3px;font-size:9px;font-weight:600}}

.link-ver{{color:var(--eu);text-decoration:none;font-size:11px;font-weight:600}}.link-ver:hover{{text-decoration:underline}}

.res{{margin-top:18px}}.rb{{padding:12px 16px;border-radius:var(--r);margin-bottom:8px}}
.rb strong{{font-size:12px;display:block;margin-bottom:5px}}
.rl{{font-size:11px;line-height:2.1}}.rl a{{font-weight:500;text-decoration:none}}.rl a:hover{{text-decoration:underline}}
.r-eu{{background:var(--eubg);border:1px solid var(--eubd)}}.r-eu a{{color:var(--eu)}}
.r-es{{background:var(--esbg);border:1px solid var(--esbd)}}.r-es a{{color:var(--es)}}
.r-eus{{background:var(--eusbg);border:1px solid var(--eusbd)}}.r-eus a{{color:var(--eus)}}
.ftr{{text-align:center;margin-top:18px;font-size:10px;color:var(--tx3);padding-bottom:16px}}
.nr{{text-align:center;padding:30px;color:var(--tx3);font-size:13px;display:none}}
@media(max-width:700px){{.sts{{grid-template-columns:repeat(3,1fr)}}.tabs{{flex-wrap:wrap}}.hdr{{flex-direction:column;gap:8px}}}}
</style></head>
<body>
<div class="ctn">
    <div class="hdr"><div class="hdr-l"><div class="hdr-ico">FR</div><div><h1>Funding Radar</h1><div class="hdr-sub">Bilbao · Misión Climática · Neutralidad 2030</div></div></div><div class="hdr-dt">{datetime.now().strftime('%d/%m/%Y %H:%M')}</div></div>

    <div class="sts">
        <div class="st"><div class="st-n">{len(all_calls)}</div><div class="st-l">Total</div></div>
        <div class="st"><div class="st-n" style="color:var(--red)">{len(new_calls)}</div><div class="st-l">Nuevas</div></div>
        <div class="st"><div class="st-n" style="color:var(--eu)">{len(eu_calls)}</div><div class="st-l">Europa</div></div>
        <div class="st"><div class="st-n" style="color:var(--es)">{len(es_calls)}</div><div class="st-l">España</div></div>
        <div class="st"><div class="st-n" style="color:var(--eus)">{len(eus_calls)}</div><div class="st-l">Euskadi</div></div>
        <div class="st"><div class="st-n" style="color:var(--grn)">{muy_alta}</div><div class="st-l">Muy Alta</div></div>
    </div>

    <div class="tabs">
        <button class="tab t-all active" onclick="fS('all',this)">Todas <span class="tc">{len(all_calls)}</span></button>
        <button class="tab t-eu" onclick="fS('eu',this)">🇪🇺 Europa <span class="tc">{len(eu_calls)}</span></button>
        <button class="tab t-es" onclick="fS('es',this)">🇪🇸 España <span class="tc">{len(es_calls)}</span></button>
        <button class="tab t-eus" onclick="fS('eus',this)">🟢 Euskadi <span class="tc">{len(eus_calls)}</span></button>
        <button class="tab t-new" onclick="fS('new',this)">🆕 Nuevas <span class="tc">{len(new_calls)}</span></button>
    </div>

    <div class="tbar">
        <input class="srch" type="text" placeholder="Buscar por título, programa, tema..." oninput="fTx(this.value)">
        <button class="fb active" onclick="fR('all',this)">Todas</button>
        <button class="fb" onclick="fR('muy-alta',this)">🔴 Muy Alta</button>
        <button class="fb" onclick="fR('alta',this)">🔵 Alta+</button>
    </div>

    <div class="xlbar">📊 <a href="resultados_convocatorias.xlsx">Descargar Excel con fichas detalladas</a></div>

    {new_section}

    <table class="calls-table" id="mt"><thead><tr class="table-head"><th>Convocatoria</th><th class="th-dl">Deadline</th><th class="th-lk"></th></tr></thead>
    <tbody id="cb">{all_rows}</tbody></table>
    <div class="nr" id="nr">No se encontraron convocatorias con esos filtros</div>

    <div class="res">
        <div class="rb r-eu"><strong>🇪🇺 Europa — Programas y convocatorias</strong><div class="rl">
            <a href="https://ec.europa.eu/info/funding-tenders/opportunities/portal/screen/opportunities/calls-for-proposals" target="_blank">Portal EU</a> ·
            <a href="https://netzerocities.eu" target="_blank">NetZeroCities</a> ·
            <a href="https://urbact.eu/calls-for-proposals" target="_blank">URBACT</a> ·
            <a href="https://interreg-sudoe.eu" target="_blank">INTERREG SUDOE</a> ·
            <a href="https://www.atlanticarea.eu" target="_blank">INTERREG Atlantic</a> ·
            <a href="https://cinea.ec.europa.eu/programmes/life_en" target="_blank">LIFE</a> ·
            <a href="https://www.climate-kic.org/programmes/" target="_blank">EIT Climate-KIC</a> ·
            <a href="https://www.eiturbanmobility.eu/calls/" target="_blank">EIT Urban Mobility</a> ·
            <a href="https://new-european-bauhaus.europa.eu/get-involved/funding-opportunities_en" target="_blank">NEB</a> ·
            <a href="https://www.urban-initiative.eu/calls-for-proposals" target="_blank">EUI-IA</a> ·
            <a href="https://www.eib.org/en/products/advising/elena/index.htm" target="_blank">BEI ELENA</a> ·
            <a href="https://advisory.eib.org/about/circular-city-centre.htm" target="_blank">EIB Circular</a> ·
            <a href="https://www.c40.org/" target="_blank">C40</a> ·
            <a href="https://ted.europa.eu/en/advanced-search" target="_blank">TED</a> ·
            <a href="https://cordis.europa.eu" target="_blank">CORDIS</a>
        </div></div>
        <div class="rb r-es"><strong>🇪🇸 España — Ayudas nacionales</strong><div class="rl">
            <a href="https://ayudasenergiaidae.es/programas-ayudas-abiertas/" target="_blank">IDAE</a> ·
            <a href="https://www.cdti.es/convocatorias" target="_blank">CDTI</a> ·
            <a href="https://fundacion-biodiversidad.es/convocatorias/" target="_blank">F. Biodiversidad</a> ·
            <a href="https://www.miteco.gob.es/es/ministerio/servicios/ayudas-subvenciones/" target="_blank">MITECO</a> ·
            <a href="https://www.pap.hacienda.gob.es/bdnstrans/GE/es/convocatorias" target="_blank">BDNS</a> ·
            <a href="https://femp-fondos-europa.es/convocatorias/" target="_blank">FEMP</a> ·
            <a href="https://www.redinnpulso.es/" target="_blank">Red Innpulso</a>
        </div></div>
        <div class="rb r-eus"><strong>🟢 Euskadi / Bizkaia</strong><div class="rl">
            <a href="https://www.eve.eus/programa-de-ayudas/" target="_blank">EVE</a> ·
            <a href="https://www.ihobe.eus/subvenciones" target="_blank">Ihobe</a> ·
            <a href="https://www.spri.eus/es/ayudas/" target="_blank">SPRI</a> ·
            <a href="https://www.bizkaia.eus/es/subvenciones" target="_blank">Dip. Bizkaia</a> ·
            <a href="https://www.euskadi.eus/ayudas-subvenciones-702/web01-tramite/es/" target="_blank">Gobierno Vasco</a> ·
            <a href="https://www.udalsarea21.net/" target="_blank">Udalsarea 2030</a> ·
            <a href="https://barrixe.com/" target="_blank">Barrixe</a> ·
            <a href="https://www.contratacion.euskadi.eus/webkpe00-kpeperfi/es/ac70cPublicidadWar/busquedaAnuncios?locale=es" target="_blank">KontratazioA</a>
        </div></div>
    </div>

    <div class="ftr">Funding Radar · Ayuntamiento de Bilbao · Fuentes: API SEDIA (EU) · BDNS (España) · euskadi.eus (Euskadi)</div>
</div>
<script>
let aS='all',aR='all',sT='';
function af(){{const rows=document.querySelectorAll('#cb .call-row');let v=0;rows.forEach(r=>{{let s=true;if(aS==='new')s=r.dataset.new==='1';else if(aS!=='all')s=r.dataset.source===aS;if(s&&aR!=='all'){{if(aR==='muy-alta')s=r.dataset.rel==='MUY ALTA';else if(aR==='alta')s=r.dataset.rel==='ALTA'||r.dataset.rel==='MUY ALTA'}}if(s&&sT){{const t=r.textContent.toLowerCase();s=sT.split(' ').every(w=>t.includes(w))}}r.style.display=s?'':'none';if(s)v++}});document.getElementById('nr').style.display=v===0?'block':'none';const ns=document.getElementById('new-section');if(ns)ns.style.display=(aS==='all'&&aR==='all'&&!sT)?'':'none'}}
function fS(s,b){{aS=s;document.querySelectorAll('.tabs .tab').forEach(t=>t.classList.remove('active'));b.classList.add('active');af()}}
function fR(r,b){{aR=r;document.querySelectorAll('.fb').forEach(x=>x.classList.remove('active'));b.classList.add('active');af()}}
function fTx(v){{sT=v.toLowerCase().trim();af()}}
</script></body></html>"""

    with open(CONFIG["output_html"], "w", encoding="utf-8") as f:
        f.write(html)
    return html

# ──────────────────────────────────────────────
# EMAIL
# ──────────────────────────────────────────────

def send_email(new_calls, all_calls):
    if not CONFIG["email_to"] or not CONFIG["smtp_user"]:
        print("\n📧 Email no configurado.")
        return
    if not new_calls:
        return

    subject = f"EU Funding Radar: {len(new_calls)} nuevas — {datetime.now().strftime('%d/%m/%Y')}"
    items = ""
    for c in sorted(new_calls.values(), key=lambda x: x.get("deadline", "9999")):
        items += f'<div style="background:#F8FAFC;border:1px solid #E2E8F0;border-radius:8px;padding:12px;margin-bottom:8px"><strong>{c["title"][:100]}</strong><br><span style="font-size:11px;color:#64748B">{c["id"]}</span><br><span style="font-size:12px;color:#475569">{c["description"][:150]}</span><br><a href="{c["url"]}" style="color:#0057B7;font-size:12px">Ver en portal</a></div>'

    body = f'<div style="font-family:sans-serif;max-width:600px;margin:0 auto"><div style="background:#0C1220;color:white;padding:20px;border-radius:12px 12px 0 0"><h1 style="font-size:18px;margin:0">EU Funding Radar</h1><p style="font-size:12px;color:#94A3B8;margin:4px 0 0">Bilbao · {datetime.now().strftime("%d/%m/%Y")}</p></div><div style="padding:20px;background:white;border:1px solid #E2E8F0;border-radius:0 0 12px 12px"><p style="margin-bottom:16px"><strong style="color:#DC2626">{len(new_calls)} convocatorias nuevas</strong></p>{items}</div></div>'

    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = CONFIG["email_from"]
        msg["To"] = CONFIG["email_to"]
        msg.attach(MIMEText(body, "html", "utf-8"))
        with smtplib.SMTP(CONFIG["smtp_server"], CONFIG["smtp_port"]) as s:
            s.starttls()
            s.login(CONFIG["smtp_user"], CONFIG["smtp_pass"])
            s.sendmail(CONFIG["email_from"], CONFIG["email_to"].split(","), msg.as_string())
        print(f"\n📧 Email enviado a {CONFIG['email_to']}")
    except Exception as e:
        print(f"\n⚠️  Error email: {e}")


# ──────────────────────────────────────────────
# BDNS — BASE DE DATOS NACIONAL DE SUBVENCIONES
# ──────────────────────────────────────────────

BDNS_KEYWORDS = [
    "eficiencia energetica",
    "energia renovable",
    "movilidad sostenible",
    "economia circular",
    "rehabilitacion energetica",
    "cambio climatico",
    "biodiversidad restauracion",
    "transicion energetica",
    "vehiculo electrico",
    "residuos urbanos",
    "ciudad inteligente",
    "descarbonizacion",
    "infraestructura verde",
    "bajas emisiones",
    "renovable autoconsumo",
    "edificio energia",
]

# Regiones relevantes para Bilbao
BDNS_REGIONES_OK = [
    "es21", "pais vasco", "euskadi", "bizkaia", "vizcaya", "bilbao",
    "alava", "araba", "gipuzkoa", "guipuzcoa",
    "es - todas", "todo el territorio", "nacional",
]

# Palabras clave para filtrar tematica relevante
BDNS_TEMAS_OK = [
    "energia", "renovable", "climatico", "clima", "eficiencia", "movilidad",
    "rehabilitacion", "residuo", "emision", "descarbonizacion", "urbano", "urbana",
    "sostenible", "sostenibilidad", "medioambient", "ambiental", "biodiversidad",
    "circular", "verde", "electrico", "hidrogeno", "transicion", "autoconsumo",
    "edificio", "vivienda", "transporte", "infraestructura", "agua", "inundacion",
    "smart", "inteligente", "digital", "innovacion", "investigacion", "i+d",
    "municipio", "ayuntamiento", "entidad local", "corporacion local",
]


def fetch_bdns_detail(num_conv):
    """Obtiene el detalle de una convocatoria BDNS por su numero"""
    try:
        url = f"https://www.infosubvenciones.es/bdnstrans/api/convocatorias?numConv={num_conv}&vpd=GE"
        req = urllib.request.Request(url, headers={
            "Accept": "application/json",
            "User-Agent": "EU-Funding-Radar/1.0"
        })
        with urllib.request.urlopen(req, timeout=20) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except:
        return None


def is_region_relevant(detail):
    """Comprueba si la convocatoria es de ambito Pais Vasco o nacional"""
    regiones = detail.get("regiones", [])

    # Si no tiene regiones, comprobar si es AGE (nacional)
    if not regiones:
        organo = detail.get("organo", {})
        nivel1 = organo.get("nivel1", "").upper()
        if nivel1 in ["ESTATAL", "AGE", "MINISTERIO"]:
            return True
        return False

    for r in regiones:
        desc = r.get("descripcion", "").lower()
        for ok in BDNS_REGIONES_OK:
            if ok in desc:
                return True

    # Si tiene muchas regiones (>10), probablemente es nacional
    if len(regiones) > 10:
        return True

    return False


def is_tema_relevant(detail):
    """Comprueba si la tematica es relevante para clima/energia/sostenibilidad"""
    titulo = detail.get("descripcion", "").lower()
    finalidad = detail.get("descripcionFinalidad", "").lower()
    bases = detail.get("descripcionBasesReguladoras", "").lower()
    fondos = " ".join([f.get("descripcion", "") for f in detail.get("fondos", [])]).lower()
    text = f"{titulo} {finalidad} {bases} {fondos}"

    for kw in BDNS_TEMAS_OK:
        if kw in text:
            return True
    return False


def fetch_bdns_calls():
    """Consulta las ultimas 200 convocatorias de la BDNS y filtra por region + tema"""
    today = datetime.now(timezone.utc)

    print(f"\n🇪🇸 BDNS -- Base de Datos Nacional de Subvenciones")
    print(f"{'='*50}")

    # Paso 1: Recoger las ultimas 500 convocatorias con pre-filtro
    all_nums = []
    prefiltro_skip = 0
    for page in range(10):  # 10 paginas x 50 = 500
        try:
            url = f"https://www.infosubvenciones.es/bdnstrans/api/convocatorias/busqueda?page={page}&pageSize=50&vpd=GE"
            req = urllib.request.Request(url, headers={
                "Accept": "application/json",
                "User-Agent": "EU-Funding-Radar/1.0"
            })
            with urllib.request.urlopen(req, timeout=30) as resp:
                data = json.loads(resp.read().decode("utf-8"))
            content = data.get("content", [])
            for conv in content:
                num = str(conv.get("numeroConvocatoria", ""))
                if not num:
                    continue
                # Pre-filtro rapido: nivel1 + nivel2 + descripcion
                nivel1 = (conv.get("nivel1", "") or "").upper()
                nivel2 = (conv.get("nivel2", "") or "").lower()
                desc = (conv.get("descripcion", "") or "").lower()

                # Descartar rapido si es local de otra region
                es_local = nivel1 == "LOCAL"
                es_euskadi_local = any(kw in nivel2 for kw in ["bilbao", "bizkaia", "vizcaya", "vitoria", "gasteiz",
                    "donostia", "san sebastian", "gipuzkoa", "guipuzcoa", "alava", "araba",
                    "euskadi", "pais vasco", "gobierno vasco", "diputacion foral"])
                es_estatal = nivel1 in ["ESTATAL", "AGE"]
                es_autonomico = nivel1 == "AUTONOMICO" or nivel1 == "AUTONÓMICO"
                es_euskadi_auto = es_autonomico and any(kw in nivel2 for kw in ["euskadi", "pais vasco", "gobierno vasco", "eve", "ihobe", "spri"])
                es_otros_nacional = nivel1 in ["OTROS", "UNIVERSIDAD"]

                # Solo pasar: locales de Euskadi, estatales, autonomicas vascas
                if es_local and not es_euskadi_local:
                    prefiltro_skip += 1
                    continue
                if es_autonomico and not es_euskadi_auto:
                    prefiltro_skip += 1
                    continue

                all_nums.append(num)
        except Exception as e:
            print(f"  ⚠️  Error pagina {page}: {str(e)[:40]}")

    print(f"🔍 500 ultimas convocatorias escaneadas")
    print(f"   Pre-filtro: {prefiltro_skip} descartadas (local/autonomica de otra region)")
    print(f"   Candidatas para detalle: {len(all_nums)}")
    print(f"🔍 Consultando detalle...\n")

    print(f"🔍 Ultimas 500 convocatorias registradas")
    print(f"🔍 Consultando detalle y filtrando por Pais Vasco / Nacional...\n")

    bdns_calls = {}
    checked = 0
    skipped_region = 0
    skipped_tema = 0
    skipped_closed = 0
    open_count = 0

    for num_conv in all_nums:
        checked += 1
        if checked % 50 == 0:
            print(f"  ... {checked}/{len(all_nums)} consultadas | {open_count} relevantes | {skipped_region} fuera de region | {skipped_tema} tema no relevante | {skipped_closed} cerradas", flush=True)

        detail = fetch_bdns_detail(num_conv)
        if not detail:
            continue

        # Filtro 1: Region (Pais Vasco o Nacional)
        if not is_region_relevant(detail):
            skipped_region += 1
            continue

        # Filtro 2: Tema relevante
        if not is_tema_relevant(detail):
            skipped_tema += 1
            continue

        # Filtro 3: Plazo abierto
        fecha_fin = detail.get("fechaFinSolicitud", "")
        abierto = detail.get("abierto", False)
        deadline_str = ""

        if fecha_fin:
            try:
                dt = datetime.strptime(fecha_fin, "%Y-%m-%d").replace(tzinfo=timezone.utc)
                if dt < today:
                    skipped_closed += 1
                    continue
                deadline_str = dt.strftime("%d/%m/%Y")
            except:
                pass

        if not abierto and not deadline_str:
            skipped_closed += 1
            continue

        # Extraer datos
        titulo = detail.get("descripcion", "Sin titulo")
        organo = detail.get("organo", {})
        organismo = organo.get("nivel2", organo.get("nivel1", ""))
        presupuesto = detail.get("presupuestoTotal", "")

        regiones = detail.get("regiones", [])
        regiones_str = ", ".join([r.get("descripcion", "") for r in regiones[:3]])

        beneficiarios = detail.get("tiposBeneficiarios", [])
        benef_str = ", ".join([b.get("descripcion", "") for b in beneficiarios[:2]])

        fondos = detail.get("fondos", [])
        fondos_str = ", ".join([f.get("descripcion", "") for f in fondos[:2]])

        url_conv = f"https://www.infosubvenciones.es/bdnstrans/GE/es/convocatoria/{num_conv}"

        # Relevancia
        text_lower = f"{titulo} {organismo} {regiones_str} {benef_str}".lower()
        relevance = "MEDIA"
        muy_alta_kw = ["municipio", "ayuntamiento", "entidad local", "corporacion local",
                       "bilbao", "euskadi", "pais vasco", "bizkaia", "vizcaya"]
        alta_kw = ["energia", "renovable", "climatico", "clima", "eficiencia", "movilidad",
                   "rehabilitacion", "residuo", "emision", "descarbonizacion", "urbano", "urbana"]
        for kw in muy_alta_kw:
            if kw in text_lower:
                relevance = "MUY ALTA"
                break
        if relevance != "MUY ALTA":
            for kw in alta_kw:
                if kw in text_lower:
                    relevance = "ALTA"
                    break

        bdns_id = f"BDNS-{num_conv}"
        bdns_calls[bdns_id] = {
            "id": bdns_id,
            "title": titulo[:200],
            "description": f"{benef_str}. {fondos_str}. Regiones: {regiones_str}".strip(". "),
            "status": "Open" if abierto else "Open",
            "deadline": deadline_str,
            "url": url_conv,
            "programme": organismo[:60] if organismo else "BDNS",
            "budget": f"{presupuesto:,.2f} EUR" if isinstance(presupuesto, (int, float)) and presupuesto > 0 else "",
            "action_type": "Subvencion Nacional",
            "call_id": f"BDNS {num_conv}",
            "tags": fondos_str,
            "source": "BDNS",
            "relevance_level": relevance,
        }
        open_count += 1

    print(f"\n📊 BDNS resumen:")
    print(f"   Consultadas: {checked}")
    print(f"   Fuera de region: {skipped_region}")
    print(f"   Tema no relevante: {skipped_tema}")
    print(f"   Cerradas/sin plazo: {skipped_closed}")
    print(f"   ✅ Relevantes abiertas: {open_count}")
    return bdns_calls


# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────

# ──────────────────────────────────────────────
# KONTRATAZIOA - LICITACIONES + AYUDAS EUSKADI
# ──────────────────────────────────────────────

EUSKADI_TEMAS_OK = [
    "energia", "renovable", "climatico", "clima", "eficiencia", "movilidad",
    "rehabilitacion", "residuo", "emision", "descarbonizacion", "urbano", "urbana",
    "sostenible", "sostenibilidad", "medioambient", "ambiental", "biodiversidad",
    "circular", "verde", "electrico", "hidrogeno", "transicion", "autoconsumo",
    "edificio", "vivienda", "transporte", "infraestructura", "agua", "inundacion",
    "smart", "inteligente", "digital", "innovacion", "alumbrado", "iluminacion",
    "parque", "jardin", "arbolado", "limpieza", "recogida", "contenedor",
    "bicicleta", "carril bici", "peatonal", "calmado trafico", "zona baja",
    "aislamiento", "cubierta", "fachada", "calefaccion", "caldera",
    "fotovoltaica", "solar", "eolica", "biomasa", "geotermia",
    "saneamiento", "depuracion", "pluvial", "drenaje",
]

# Busquedas tematicas para el buscador de euskadi.eus
EUSKADI_SEARCH_QUERIES = [
    # (tipo_tramite, keyword)
    ("ayuda_subvencion", "energia"),
    ("ayuda_subvencion", "clima"),
    ("ayuda_subvencion", "movilidad sostenible"),
    ("ayuda_subvencion", "rehabilitacion edificios"),
    ("ayuda_subvencion", "medio ambiente"),
    ("ayuda_subvencion", "residuos circular"),
    ("ayuda_subvencion", "renovable"),
    ("ayuda_subvencion", "biodiversidad"),
    ("ayuda_subvencion", "innovacion"),
    ("ayuda_subvencion", "transporte"),
    ("ayuda_subvencion", "agua"),
    ("ayuda_subvencion", "digital"),
    ("anuncio_contratacion", "energia"),
    ("anuncio_contratacion", "movilidad"),
    ("anuncio_contratacion", "residuos"),
    ("anuncio_contratacion", "agua saneamiento"),
    ("anuncio_contratacion", "rehabilitacion"),
    ("anuncio_contratacion", "alumbrado"),
    ("anuncio_contratacion", "medio ambiente"),
    ("anuncio_contratacion", "transporte sostenible"),
]


def fetch_kontratazioa_calls():
    """Consulta licitaciones y ayudas de Euskadi via API de euskadi.eus"""
    today = datetime.now(timezone.utc)

    print(f"\n🟢 Euskadi -- Contrataciones y Ayudas del Sector Publico Vasco")
    print(f"{'='*50}")

    eus_calls = {}

    # ──── ESTRATEGIA 1: API eventos administrativos ────
    # Patron: api.euskadi.eus/{domain}/{version}/{resource}
    api_endpoints = [
        # Eventos administrativos (contrataciones + ayudas)
        "https://api.euskadi.eus/administration/events/v1.0/events/byType/anuncio_contratacion?_page=1&_pageSize=100",
        "https://api.euskadi.eus/administration/v1.0/events/byType/anuncio_contratacion?_page=1&_pageSize=100",
        # Subvenciones / ayudas
        "https://api.euskadi.eus/administration/events/v1.0/events/byType/ayuda_subvencion?_page=1&_pageSize=100",
        "https://api.euskadi.eus/administration/v1.0/events/byType/ayuda_subvencion?_page=1&_pageSize=100",
    ]

    api_data_found = False
    for url in api_endpoints:
        try:
            req = urllib.request.Request(url, headers={
                "Accept": "application/json",
                "User-Agent": "EU-Funding-Radar/1.0"
            })
            with urllib.request.urlopen(req, timeout=20) as resp:
                raw = resp.read().decode("utf-8")
                data = json.loads(raw)
                items = []
                if isinstance(data, list):
                    items = data
                elif isinstance(data, dict):
                    for key in ["items", "content", "results", "data"]:
                        if key in data and isinstance(data[key], list):
                            items = data[key]
                            break

                tipo = "contratacion" if "contratacion" in url else "ayuda"
                print(f"  ✓ API eventos ({tipo}): {len(items)} items")
                api_data_found = True

                for item in items:
                    parsed = parse_euskadi_item(item, today)
                    if parsed:
                        eus_calls[parsed["id"]] = parsed

        except Exception as e:
            err = str(e)[:60]
            tipo = "contratacion" if "contratacion" in url else "ayuda"
            print(f"  ⚠️  API eventos ({tipo}): {err}")

    # ──── ESTRATEGIA 2: Buscador euskadi.eus (contenido web) ────
    if not api_data_found:
        print(f"  API eventos no disponible. Usando buscador web...")

    print(f"🔍 Buscando en euskadi.eus ({len(EUSKADI_SEARCH_QUERIES)} busquedas)...")
    seen_urls = set()
    for i, (tipo, kw) in enumerate(EUSKADI_SEARCH_QUERIES, 1):
        try:
            encoded_kw = urllib.parse.quote(kw)
            # URL del buscador de tramites de euskadi.eus
            search_url = (
                f"https://www.euskadi.eus/gobierno-vasco/tramites-servicios/"
                f"?r01kQry=tT:{tipo};tC:{encoded_kw}"
            )
            req = urllib.request.Request(search_url, headers={
                "Accept": "text/html, */*",
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            })
            with urllib.request.urlopen(req, timeout=15) as resp:
                raw_bytes = resp.read()
                # euskadi.eus usa ISO-8859-1 / Latin-1
                ct = resp.headers.get("Content-Type", "")
                if "iso-8859" in ct.lower() or "latin" in ct.lower():
                    raw = raw_bytes.decode("iso-8859-1")
                else:
                    try:
                        raw = raw_bytes.decode("utf-8")
                    except UnicodeDecodeError:
                        raw = raw_bytes.decode("iso-8859-1")

                # Patron exacto: <em class="r01srItemDocName"><a href="/ayuda_subvencion/...">titulo</a></em>
                results = re.findall(
                    r'<a\s+href="(/(?:ayuda_subvencion|anuncio_contratacion)/[^"]+)"[^>]*>([^<]+)</a>',
                    raw, re.IGNORECASE
                )

                new = 0
                for url_path, title_raw in results:
                    title_clean = title_raw.strip()
                    if len(title_clean) < 15:
                        continue

                    full_url = f"https://www.euskadi.eus{url_path}"
                    if full_url in seen_urls:
                        continue
                    seen_urls.add(full_url)

                    text_lower = title_clean.lower()

                    # Filtro tematico
                    if not any(t in text_lower for t in EUSKADI_TEMAS_OK):
                        continue

                    eus_id = f"EUS-{abs(hash(full_url)) % 100000:05d}"
                    if eus_id in eus_calls:
                        continue

                    relevance = "MEDIA"
                    for mk in ["bilbao", "ayuntamiento", "municipio"]:
                        if mk in text_lower:
                            relevance = "MUY ALTA"
                            break
                    if relevance != "MUY ALTA":
                        for ak in ["energia", "renovable", "clima", "eficiencia", "movilidad",
                                   "rehabilitacion", "residuo", "emision", "sostenible",
                                   "descarbonizacion"]:
                            if ak in text_lower:
                                relevance = "ALTA"
                                break

                    if "ayuda_subvencion" in url_path:
                        tipo_accion = "Ayuda/Subvencion Euskadi"
                    else:
                        tipo_accion = "Licitacion Euskadi"

                    eus_calls[eus_id] = {
                        "id": eus_id,
                        "title": title_clean[:200],
                        "description": "Gobierno Vasco / Sector Publico Euskadi",
                        "status": "Open",
                        "deadline": "",
                        "url": full_url,
                        "programme": "Euskadi",
                        "budget": "",
                        "action_type": tipo_accion,
                        "call_id": "",
                        "tags": kw,
                        "source": "KontratazioA",
                        "relevance_level": relevance,
                    }
                    new += 1

                tipo_label = "ayudas" if "ayuda" in tipo else "licitaciones"
                print(f"  [{i}/{len(EUSKADI_SEARCH_QUERIES)}] {tipo_label}: {kw}... "
                      f"{len(results)} resultados, {new} nuevas relevantes")
            time.sleep(0.3)
        except Exception as e:
            print(f"  [{i}/{len(EUSKADI_SEARCH_QUERIES)}] {kw}... ⚠️ {str(e)[:40]}")

    # ──── ESTRATEGIA 3: JSON datasets de contrataciones ────
    # Los datasets se publican periodicamente en Open Data Euskadi
    json_urls = [
        "https://opendata.euskadi.eus/contenidos/ds_contrataciones/contrataciones_702/opendata/contrataciones.json",
    ]
    for jurl in json_urls:
        try:
            req = urllib.request.Request(jurl, headers={
                "Accept": "application/json",
                "User-Agent": "EU-Funding-Radar/1.0"
            })
            with urllib.request.urlopen(req, timeout=30) as resp:
                raw = resp.read().decode("utf-8", errors="replace")
                data = json.loads(raw)
                if isinstance(data, list):
                    items = data
                elif isinstance(data, dict):
                    items = data.get("items", data.get("contrataciones", []))
                else:
                    items = []

                parsed_count = 0
                for item in items:
                    parsed = parse_euskadi_item(item, today)
                    if parsed and parsed["id"] not in eus_calls:
                        eus_calls[parsed["id"]] = parsed
                        parsed_count += 1

                if parsed_count > 0:
                    print(f"  ✓ Dataset JSON: {parsed_count} licitaciones relevantes")
        except Exception as e:
            pass  # Silencioso, dataset puede no existir

    print(f"\n📊 Euskadi resumen:")
    print(f"   ✅ Resultados relevantes: {len(eus_calls)}")
    if not eus_calls:
        print(f"   ℹ️  Sin resultados automaticos. Consultar manualmente:")
        print(f"      https://www.contratacion.euskadi.eus")
        print(f"      https://www.euskadi.eus/gobierno-vasco/tramites-servicios/")
    return eus_calls


def parse_euskadi_item(item, today):
    """Parsea un item de la API de euskadi.eus en formato estandar"""
    if not isinstance(item, dict):
        return None

    # Extraer titulo (diferentes formatos posibles)
    titulo = ""
    for key in ["title", "titulo", "nombre", "descripcion", "objectContract", "text"]:
        val = item.get(key, "")
        if val:
            if isinstance(val, dict):
                val = val.get("es", val.get("eu", str(val)))
            titulo = str(val)
            break

    if not titulo or len(titulo) < 10:
        return None

    text_lower = titulo.lower()

    # Filtro tematico
    if not any(kw in text_lower for kw in EUSKADI_TEMAS_OK):
        return None

    # ID
    item_id = str(item.get("id", item.get("expedientNumber", item.get("code", hash(titulo) % 100000))))
    eus_id = f"EUS-{item_id}"

    # URL
    url_conv = item.get("url", item.get("link", item.get("publicUrl", "")))
    if not url_conv:
        url_conv = "https://www.contratacion.euskadi.eus"

    # Relevancia
    relevance = "MEDIA"
    for mk in ["bilbao", "ayuntamiento", "municipio"]:
        if mk in text_lower:
            relevance = "MUY ALTA"
            break
    if relevance != "MUY ALTA":
        for ak in ["energia", "renovable", "clima", "eficiencia", "movilidad",
                   "rehabilitacion", "residuo", "emision", "sostenible"]:
            if ak in text_lower:
                relevance = "ALTA"
                break

    # Organismo
    organismo = ""
    for key in ["contractingAuthorityName", "poderAdjudicador", "buyerName", "organismo"]:
        val = item.get(key, "")
        if val:
            if isinstance(val, dict):
                val = val.get("es", str(val))
            organismo = str(val)[:60]
            break

    return {
        "id": eus_id,
        "title": titulo[:200],
        "description": organismo or "Sector Publico Euskadi",
        "status": "Open",
        "deadline": "",
        "url": url_conv,
        "programme": organismo[:40] if organismo else "Euskadi",
        "budget": "",
        "action_type": "Contratacion/Ayuda Euskadi",
        "call_id": item_id,
        "tags": "",
        "source": "KontratazioA",
        "relevance_level": relevance,
    }


def main():
    all_calls = fetch_all_calls()

    # Intentar BDNS (fuentes nacionales espanolas)
    try:
        bdns_calls = fetch_bdns_calls()
        if bdns_calls:
            all_calls.update(bdns_calls)
            print(f"📊 Total combinado (EU + BDNS): {len(all_calls)}")
    except Exception as e:
        print(f"\n⚠️  BDNS no disponible: {e}")
        print("   Continuando solo con convocatorias europeas...")

    # Intentar KontratazioA (licitaciones Euskadi)
    try:
        eus_calls = fetch_kontratazioa_calls()
        if eus_calls:
            all_calls.update(eus_calls)
            print(f"📊 Total combinado (EU + BDNS + Euskadi): {len(all_calls)}")
    except Exception as e:
        print(f"\n⚠️  KontratazioA no disponible: {e}")
        print("   Continuando sin licitaciones vascas...")

    if not all_calls:
        print("\n❌ No se encontraron convocatorias.")
        return 1

    seen = load_seen()
    new_calls = {k: v for k, v in all_calls.items() if k not in seen}
    print(f"🆕 Nuevas desde ultima ejecucion: {len(new_calls)}")

    # Guardar JSON
    with open(CONFIG["output_file"], "w", encoding="utf-8") as f:
        json.dump(list(all_calls.values()), f, ensure_ascii=False, indent=2)

    # Generar HTML
    generate_html(all_calls, new_calls)

    # Generar Excel
    generate_excel(all_calls, new_calls)

    # Email
    if new_calls:
        send_email(new_calls, all_calls)

    # Actualizar vistos
    seen.update({k: datetime.now(timezone.utc).isoformat() for k in all_calls})
    save_seen(seen)

    print(f"\n{'='*50}")
    print(f"✅ COMPLETADO")
    print(f"   📊 {len(all_calls)} convocatorias")
    print(f"   🆕 {len(new_calls)} nuevas")
    print(f"   📄 HTML: {CONFIG['output_html']}")
    print(f"   📊 Excel: {CONFIG['output_excel']}")
    print(f"{'='*50}\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
